"""
compare_preload.py
==================
Standalone module — run independently from main.py.

Reads:
  1. A Tools Report Excel  (*Tools Report*.xlsx)  from TOOL_REPORT_DIR
  2. A Preload Excel file                         from PRELOAD_DIR

Compares all tools, consumables, and expendables against the preload,
matching on Part Number.

Output Excel:
  Sheet 0 – Summary
  Sheet 1 – Support Equipment
  Sheet 2 – Consumable Materials
  Sheet 3 – Expendables-Parts

Usage:
    python compare_preload.py
"""

import os
import re
import sys
import glob
import logging
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import compare_config as cfg

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

# ── Colours ────────────────────────────────────────────────────────────────────
HDR_TOOL_BG = "1F4E79"
HDR_CONS_BG = "375623"
HDR_EXP_BG  = "7B3F00"
HDR_FG      = "FFFFFF"
OK_BG       = "E2EFDA"   # green  — in preload + available
UNAVAIL_BG  = "FFF2CC"   # yellow — in preload but unavailable
MISS_BG     = "FFE0E0"   # red    — not in preload
MISS_FG     = "C00000"
ALT_BG      = "F2F2F2"   # grey   — STD tools (no PN, cannot match)

THIN   = Side(style="thin", color="AAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ── Style helpers ──────────────────────────────────────────────────────────────

def _hdr(cell, bg=HDR_TOOL_BG):
    cell.font      = Font(name="Arial", bold=True, color=HDR_FG, size=10)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = BORDER


def _dat(cell, bg=None, bold=False, color="000000"):
    cell.font      = Font(name="Arial", size=9, bold=bold, color=color)
    cell.fill      = PatternFill("solid", start_color=bg) if bg else PatternFill()
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    cell.border    = BORDER


def _widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ── File resolution ────────────────────────────────────────────────────────────

def _latest_file(directory: str, pattern: str) -> Optional[str]:
    matches = glob.glob(os.path.join(directory, pattern))
    return max(matches, key=os.path.getmtime) if matches else None


def _resolve_tool_report() -> str:
    if cfg.TOOL_REPORT_FILE:
        if not os.path.isfile(cfg.TOOL_REPORT_FILE):
            logger.error("TOOL_REPORT_FILE not found: %s", cfg.TOOL_REPORT_FILE)
            sys.exit(1)
        return cfg.TOOL_REPORT_FILE
    path = _latest_file(cfg.TOOL_REPORT_DIR, "*Tools Report*.xlsx")
    if not path:
        logger.error("No Tools Report found in: %s", cfg.TOOL_REPORT_DIR)
        sys.exit(1)
    logger.info("Tool report : %s", os.path.basename(path))
    return path


def _resolve_preload() -> str:
    if cfg.PRELOAD_FILE:
        if not os.path.isfile(cfg.PRELOAD_FILE):
            logger.error("PRELOAD_FILE not found: %s", cfg.PRELOAD_FILE)
            sys.exit(1)
        return cfg.PRELOAD_FILE
    path = _latest_file(cfg.PRELOAD_DIR, "*.xlsx")
    if not path:
        logger.error("No preload file found in: %s", cfg.PRELOAD_DIR)
        sys.exit(1)
    logger.info("Preload file: %s", os.path.basename(path))
    return path


# ── Preload loader ─────────────────────────────────────────────────────────────

def _load_preload(path: str) -> dict:
    """
    Returns dict: PART_NO_UPPER -> {qty, is_tool, availability, description}
    Duplicate part numbers are summed on Qty.
    """
    df = pd.read_excel(path, header=0, dtype={"Part. No.": str})

    required = {"Part. No.", "Part Description", "Qty", "Tool", "Availability"}
    missing  = required - set(df.columns)
    if missing:
        logger.error("Preload file missing columns: %s", missing)
        sys.exit(1)

    df = df.dropna(subset=["Part. No."])
    df["Part. No."] = df["Part. No."].astype(str).str.strip().str.upper()
    df["Qty"]       = pd.to_numeric(df["Qty"], errors="coerce").fillna(0)

    lookup = {}
    for pn, grp in df.groupby("Part. No."):
        lookup[pn] = {
            "qty":          int(grp["Qty"].sum()),
            "is_tool":      bool(grp["Tool"].iloc[0]),
            "availability": str(grp["Availability"].iloc[0]),
            "description":  str(grp["Part Description"].iloc[0]),
        }

    logger.info("Preload loaded: %d unique part numbers.", len(lookup))
    return lookup


# ── PN extraction from Support Equipment sheet ─────────────────────────────────

def _extract_pn(cell_value: str) -> Optional[str]:
    """
    Extract part number from a 'Part Numbers / Effectivity' cell.
    Cell format produced by excel_writer.py:
        '✔ K52016-1'   (in stock)
        '  G45004-31'  (not in stock)
        '—'            (STD tools — no part number)
    Strips the checkmark/spaces prefix and returns the PN,
    or None if the cell is empty / dash.
    """
    if not cell_value or pd.isna(cell_value):
        return None
    cleaned = re.sub(r'^[✔\s]+', '', str(cell_value)).strip()
    if cleaned in ('', '—', 'nan'):
        return None
    return cleaned.upper()


# ── Row builders ───────────────────────────────────────────────────────────────

def _build_support_equipment(tool_report_path: str, preload: dict) -> list:
    try:
        df = pd.read_excel(tool_report_path, sheet_name="Support Equipment",
                           header=0, dtype=str)
    except Exception as e:
        logger.error("Cannot read 'Support Equipment' sheet: %s", e)
        return []

    rows = []
    seen = set()   # deduplicate by reference_id

    for _, r in df.iterrows():
        ref_id = str(r.get("Reference ID", "") or "").strip()
        if not ref_id or ref_id == "nan" or ref_id in seen:
            continue
        seen.add(ref_id)

        desc     = str(r.get("Description", "") or "").strip()
        pn_cell  = r.get("Part Numbers / Effectivity", "")
        amm_ref  = str(r.get("Found In Tasks", "") or "").strip()

        pn = _extract_pn(pn_cell)

        if pn:
            # ── Major path: match by Part Number from Effectivity column ──
            entry       = preload.get(pn)
            in_preload  = entry is not None
            preload_qty = entry["qty"]          if in_preload else 0
            avail       = entry["availability"] if in_preload else "—"
            notes       = "" if in_preload else "Not found in preload"
            rows.append({
                "Reference ID":        ref_id,
                "Part Number":         pn,
                "Description":         desc,
                "AMM Reference":       amm_ref,
                "Is Tool":             "Yes",
                "In Preload":          "YES" if in_preload else "NO",
                "Preload Qty":         preload_qty,
                "Availability":        avail,
                "Notes":               notes,
                "_in_preload":         in_preload,
                "_avail":              avail,
                "_no_pn":              False,
            })
        else:
            # ── Fallback: STD tool with no PN — try Reference ID ──────────
            entry       = preload.get(ref_id.upper())
            in_preload  = entry is not None
            preload_qty = entry["qty"]          if in_preload else 0
            avail       = entry["availability"] if in_preload else "—"
            notes       = "No part number — matched by Reference ID"
            if not in_preload:
                notes += " | Not found in preload"
            rows.append({
                "Reference ID":        ref_id,
                "Part Number":         ref_id,   # show ref ID in PN column
                "Description":         desc,
                "AMM Reference":       amm_ref,
                "Is Tool":             "Yes",
                "In Preload":          "YES" if in_preload else "NO",
                "Preload Qty":         preload_qty,
                "Availability":        avail,
                "Notes":               notes,
                "_in_preload":         in_preload,
                "_avail":              avail,
                "_no_pn":              True,
            })

    found   = sum(1 for r in rows if r["_in_preload"])
    missing = sum(1 for r in rows if not r["_in_preload"])
    logger.info("Support Equipment : %d items | %d in preload | %d missing",
                len(rows), found, missing)
    return rows


def _build_consumables(tool_report_path: str, preload: dict) -> list:
    try:
        df = pd.read_excel(tool_report_path, sheet_name="Consumable Materials",
                           header=0, dtype=str)
    except Exception as e:
        logger.error("Cannot read 'Consumable Materials' sheet: %s", e)
        return []

    rows = []
    seen = set()

    for _, r in df.iterrows():
        ref_id = str(r.get("Reference ID", "") or "").strip().upper()
        if not ref_id or ref_id == "NAN" or ref_id in seen:
            continue
        seen.add(ref_id)

        desc    = str(r.get("Description", "")   or "").strip()
        spec    = str(r.get("Specification", "") or "").strip()
        amm_ref = str(r.get("Found In Tasks", "") or "").strip()

        # Consumable reference IDs (e.g. G50736) are used directly as the PN
        entry       = preload.get(ref_id)
        in_preload  = entry is not None
        preload_qty = entry["qty"]          if in_preload else 0
        avail       = entry["availability"] if in_preload else "—"

        rows.append({
            "Reference ID":  ref_id,
            "Part Number":   ref_id,
            "Description":   desc,
            "Specification": spec,
            "AMM Reference": amm_ref,
            "Is Tool":       "No",
            "In Preload":    "YES" if in_preload else "NO",
            "Preload Qty":   preload_qty,
            "Availability":  avail,
            "Notes":         "" if in_preload else "Not found in preload",
            "_in_preload":   in_preload,
            "_avail":        avail,
            "_no_pn":        False,
        })

    found   = sum(1 for r in rows if r["_in_preload"])
    missing = sum(1 for r in rows if not r["_in_preload"])
    logger.info("Consumable Materials: %d items | %d in preload | %d missing",
                len(rows), found, missing)
    return rows


def _build_expendables(tool_report_path: str, preload: dict) -> list:
    try:
        df = pd.read_excel(tool_report_path, sheet_name="Expendables-Parts",
                           header=0, dtype=str)
    except Exception as e:
        logger.error("Cannot read 'Expendables-Parts' sheet: %s", e)
        return []

    rows = []

    for _, r in df.iterrows():
        pn = str(r.get("Part Number", "") or "").strip().upper()
        if not pn or pn in ("NAN", "NOT RESOLVED"):
            continue

        desc     = str(r.get("Part Description", "")    or "").strip()
        amm_item = str(r.get("AMM Item Description", "") or "").strip()
        ipd_dmc  = str(r.get("IPD DMC", "")             or "").strip()
        amm_qty  = str(r.get("Qty", "")                 or "").strip()
        src_task = str(r.get("Source Task", "")          or "").strip()

        entry       = preload.get(pn)
        in_preload  = entry is not None
        preload_qty = entry["qty"]          if in_preload else 0
        avail       = entry["availability"] if in_preload else "—"

        rows.append({
            "Part Number":   pn,
            "Description":   desc,
            "AMM Item":      amm_item,
            "IPD DMC":       ipd_dmc,
            "AMM Qty":       amm_qty,
            "Source Task":   src_task,
            "Is Tool":       "No",
            "In Preload":    "YES" if in_preload else "NO",
            "Preload Qty":   preload_qty,
            "Availability":  avail,
            "Notes":         "" if in_preload else "Not found in preload",
            "_in_preload":   in_preload,
            "_avail":        avail,
            "_no_pn":        False,
        })

    found   = sum(1 for r in rows if r["_in_preload"])
    missing = sum(1 for r in rows if not r["_in_preload"])
    logger.info("Expendables-Parts   : %d items | %d in preload | %d missing",
                len(rows), found, missing)
    return rows


# ── Row background ─────────────────────────────────────────────────────────────

def _bg(row: dict) -> str:
    '''if row["_no_pn"]:
        return ALT_BG          # grey — STD tool, no PN to match'''
    if not row["_in_preload"]:
        return MISS_BG         # red  — missing
    if row["_avail"] == "UNAVAILABLE":
        return UNAVAIL_BG      # yellow — present but unavailable
    return OK_BG               # green — present and available


# ── Sheet writers ──────────────────────────────────────────────────────────────

def _write_support_equipment(wb, rows: list) -> None:
    ws = wb.create_sheet("Support Equipment")
    ws.freeze_panes = "A2"

    headers = ["Reference ID", "Part Number", "Description",
               "AMM Reference", "Is Tool", "In Preload",
               "Preload Qty", "Availability", "Notes"]
    for col, h in enumerate(headers, 1):
        _hdr(ws.cell(row=1, column=col, value=h), bg=HDR_TOOL_BG)
    ws.row_dimensions[1].height = 28

    for row_idx, row in enumerate(rows, 2):
        bg = _bg(row)
        for col, h in enumerate(headers, 1):
            val   = row.get(h, "")
            bold  = h == "In Preload"
            color = MISS_FG if (h == "In Preload" and val == "NO" and not row["_no_pn"]) else "000000"
            _dat(ws.cell(row=row_idx, column=col, value=val),
                 bg=bg, bold=bold, color=color)

    _widths(ws, {"A": 14, "B": 18, "C": 42, "D": 45,
                 "E": 10, "F": 12, "G": 12, "H": 16, "I": 35})


def _write_consumables(wb, rows: list) -> None:
    ws = wb.create_sheet("Consumable Materials")
    ws.freeze_panes = "A2"

    headers = ["Reference ID", "Part Number", "Description", "Specification",
               "AMM Reference", "Is Tool", "In Preload",
               "Preload Qty", "Availability", "Notes"]
    for col, h in enumerate(headers, 1):
        _hdr(ws.cell(row=1, column=col, value=h), bg=HDR_CONS_BG)
    ws.row_dimensions[1].height = 28

    for row_idx, row in enumerate(rows, 2):
        bg = _bg(row)
        for col, h in enumerate(headers, 1):
            val   = row.get(h, "")
            bold  = h == "In Preload"
            color = MISS_FG if (h == "In Preload" and val == "NO") else "000000"
            _dat(ws.cell(row=row_idx, column=col, value=val),
                 bg=bg, bold=bold, color=color)

    _widths(ws, {"A": 14, "B": 14, "C": 42, "D": 25,
                 "E": 45, "F": 10, "G": 12, "H": 12, "I": 16, "J": 35})


def _write_expendables(wb, rows: list) -> None:
    ws = wb.create_sheet("Expendables-Parts")
    ws.freeze_panes = "A2"

    headers = ["Part Number", "Description", "AMM Item", "IPD DMC",
               "AMM Qty", "Source Task", "Is Tool", "In Preload",
               "Preload Qty", "Availability", "Notes"]
    for col, h in enumerate(headers, 1):
        _hdr(ws.cell(row=1, column=col, value=h), bg=HDR_EXP_BG)
    ws.row_dimensions[1].height = 28

    for row_idx, row in enumerate(rows, 2):
        bg = _bg(row)
        for col, h in enumerate(headers, 1):
            val   = row.get(h, "")
            bold  = h == "In Preload"
            color = MISS_FG if (h == "In Preload" and val == "NO") else "000000"
            _dat(ws.cell(row=row_idx, column=col, value=val),
                 bg=bg, bold=bold, color=color)

    _widths(ws, {"A": 18, "B": 35, "C": 25, "D": 35,
                 "E": 10, "F": 40, "G": 10, "H": 12, "I": 12, "J": 16, "K": 35})


def _write_summary(wb, tool_rows, cons_rows, exp_rows,
                   tool_report_name: str, preload_name: str) -> None:
    ws = wb.create_sheet("Summary", 0)

    # Title
    title = ws.cell(row=1, column=1,
                    value=f"Preload Comparison Report")
    title.font      = Font(name="Arial", bold=True, size=13, color="1F4E79")
    title.alignment = Alignment(vertical="center")
    ws.merge_cells("A1:D1")
    ws.row_dimensions[1].height = 24

    sub = ws.cell(row=2, column=1,
                  value=f"Tool Report : {tool_report_name}")
    sub.font = Font(name="Arial", size=9, italic=True, color="555555")
    ws.merge_cells("A2:D2")

    sub2 = ws.cell(row=3, column=1,
                   value=f"Preload File: {preload_name}")
    sub2.font = Font(name="Arial", size=9, italic=True, color="555555")
    ws.merge_cells("A3:D3")

    def _section(start_row, label, rows, hdr_bg):
        total   = len(rows)
        found   = sum(1 for r in rows if r["_in_preload"])
        missing = sum(1 for r in rows if not r["_in_preload"] and not r["_no_pn"])
        no_pn   = sum(1 for r in rows if r["_no_pn"])
        unavail = sum(1 for r in rows if r["_in_preload"] and r["_avail"] == "UNAVAILABLE")

        # Section header
        hc = ws.cell(row=start_row, column=1, value=label)
        hc.font      = Font(name="Arial", bold=True, size=10, color=HDR_FG)
        hc.fill      = PatternFill("solid", start_color=hdr_bg)
        hc.alignment = Alignment(vertical="center")
        hc.border    = BORDER
        ws.merge_cells(start_row=start_row, start_column=1,
                       end_row=start_row, end_column=2)
        ws.row_dimensions[start_row].height = 20

        data = [
            ("Total items",                     total,   None,       "000000"),
            ("In preload ✔",                    found,   OK_BG,      "000000"),
            ("Missing from preload ✗",          missing, MISS_BG,    MISS_FG),
            ("In preload but UNAVAILABLE",      unavail, UNAVAIL_BG, "000000"),
            ("No part number (STD tools)",      no_pn,   ALT_BG,     "555555"),
        ]
        for i, (lbl, val, bg, fg) in enumerate(data):
            r = start_row + 1 + i
            lc = ws.cell(row=r, column=1, value=lbl)
            lc.font   = Font(name="Arial", size=9, bold=True)
            lc.border = BORDER
            lc.fill   = PatternFill("solid", start_color=bg) if bg else PatternFill()
            vc = ws.cell(row=r, column=2, value=val)
            vc.font      = Font(name="Arial", size=9, color=fg)
            vc.fill      = PatternFill("solid", start_color=bg) if bg else PatternFill()
            vc.border    = BORDER
            vc.alignment = Alignment(horizontal="center")

        return start_row + len(data) + 2

    next_row = _section(5,  "Support Equipment",   tool_rows, HDR_TOOL_BG)
    next_row = _section(next_row, "Consumable Materials", cons_rows, HDR_CONS_BG)
    _section(next_row, "Expendables / Parts",  exp_rows, HDR_EXP_BG)

    _widths(ws, {"A": 34, "B": 14})


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    logger.info("=" * 60)
    logger.info("Preload Comparator")
    logger.info("=" * 60)

    tool_report_path = _resolve_tool_report()
    preload_path     = _resolve_preload()

    logger.info("Loading preload...")
    preload = _load_preload(preload_path)

    logger.info("Comparing Support Equipment...")
    tool_rows = _build_support_equipment(tool_report_path, preload)

    logger.info("Comparing Consumable Materials...")
    cons_rows = _build_consumables(tool_report_path, preload)

    logger.info("Comparing Expendables/Parts...")
    exp_rows  = _build_expendables(tool_report_path, preload)

    wb = Workbook()
    wb.remove(wb.active)

    _write_summary(wb, tool_rows, cons_rows, exp_rows,
                   os.path.basename(tool_report_path),
                   os.path.basename(preload_path))
    _write_support_equipment(wb, tool_rows)
    _write_consumables(wb, cons_rows)
    _write_expendables(wb, exp_rows)

    tool_base    = os.path.splitext(os.path.basename(tool_report_path))[0]
    preload_base = os.path.splitext(os.path.basename(preload_path))[0]
    out_name = f"{tool_base} vs {preload_base} - Comparison.xlsx"
    out_path = os.path.join(cfg.COMPARE_OUTPUT_DIR, out_name)
    os.makedirs(cfg.COMPARE_OUTPUT_DIR, exist_ok=True)
    wb.save(out_path)

    total_missing = (
        sum(1 for r in tool_rows if not r["_in_preload"] and not r["_no_pn"]) +
        sum(1 for r in cons_rows if not r["_in_preload"]) +
        sum(1 for r in exp_rows  if not r["_in_preload"])
    )

    logger.info("=" * 60)
    logger.info("Saved to   : %s", out_path)
    logger.info("Total missing from preload: %d", total_missing)
    logger.info("Done.")


if __name__ == "__main__":
    main()