"""
excel_writer.py
===============
Three-sheet Excel report:
  Sheet 1 – "Support Equipment"   (deduplicated, with stock check)
  Sheet 2 – "Consumable Materials" (deduplicated by reference_id)
  Sheet 3 – "Expendables/Parts"   (one row per AMM item × IPD figure)
"""

import os
from typing import List, Dict, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from extractor.html_parser          import TaskData, ToolEntry, ConsumableEntry, ExpendableEntry
from extractor.crawler              import NOT_FOUND_MARKER
from extractor.stock_checker        import StockChecker, StockResult, _parse_part_numbers
from extractor.consumable_list_parser import ConsumableListParser

# ── Colours ────────────────────────────────────────────────────────────────────
HEADER_BG    = "1F4E79"
HEADER_FG    = "FFFFFF"
PARENT_BG    = "D6E4F0"
REF_BG       = "EBF5FB"
ALT_BG       = "F2F2F2"
NOT_FOUND_BG = "FFE0E0"
NOT_FOUND_FG = "C00000"
STOCK_OK_BG  = "E2EFDA"
STOCK_NO_BG  = "FFF2CC"
CONS_HDR_BG  = "375623"   # dark green header for consumables
EXP_HDR_BG   = "7B3F00"   # dark brown header for expendables
EXP_ROW_BG   = "FFF8F0"   # light warm for expendable rows
EXP_ALT_BG   = "FFE8D0"

THIN   = Side(style="thin", color="AAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _hdr(cell, bg=HEADER_BG, fg=HEADER_FG):
    cell.font      = Font(name="Arial", bold=True, color=fg, size=10)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = BORDER

def _dat(cell, bg=None, bold=False, color="000000", italic=False):
    cell.font      = Font(name="Arial", size=9, bold=bold, color=color, italic=italic)
    cell.fill      = PatternFill("solid", start_color=bg) if bg else PatternFill()
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    cell.border    = BORDER

def _widths(ws, widths: Dict[str, int]):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ── Public entry point ─────────────────────────────────────────────────────────

def write_report(all_tasks: List[TaskData], output_path: str,
                 task_groups=None,
                 stock_checker: Optional[StockChecker] = None,
                 tool_list=None,
                 consumable_list: Optional[ConsumableListParser] = None) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    _write_support_equipment(wb, all_tasks, stock_checker, tool_list)
    _write_consumables(wb, all_tasks, consumable_list)
    _write_expendables(wb, all_tasks)
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)


# ── Sheet 1: Support Equipment ────────────────────────────────────────────────

def _write_support_equipment(wb, all_tasks, stock_checker, tool_list):
    ws = wb.create_sheet("Support Equipment")
    ws.freeze_panes = "A2"
    has_stock = stock_checker is not None

    headers = ["Reference ID", "Description", "Part Numbers / Effectivity",
               "Found In Tasks", "Occurrences"]
    if has_stock:
        headers += ["Matched Part #", "Station", "Store", "Location", "Qty"]

    for col, h in enumerate(headers, 1):
        _hdr(ws.cell(row=1, column=col, value=h))
    ws.row_dimensions[1].height = 28

    # Dedup by reference_id
    dedup: Dict[str, dict] = {}
    missing_dmcs = []
    for task in all_tasks:
        if task.title == NOT_FOUND_MARKER:
            missing_dmcs.append(task.dmc)
            continue
        for tool in task.tools:
            rid = tool.reference_id
            if rid not in dedup:
                dedup[rid] = {"description": tool.description,
                               "part_numbers": tool.part_numbers,
                               "task_titles": [], "tool_ref": tool}
            if tool.source_title not in dedup[rid]["task_titles"]:
                dedup[rid]["task_titles"].append(tool.source_title)

    def sort_key(r):
        return (0 if r.startswith("SPL") else 1 if r.startswith("STD") else 2, r)

    row_idx = 2
    for ref_id in sorted(dedup.keys(), key=sort_key):
        entry   = dedup[ref_id]
        tool    = entry["tool_ref"]
        titles  = "\n".join(entry["task_titles"])
        count   = len(entry["task_titles"])

        # Enrich from tool list if available
        tl = tool_list.lookup(ref_id) if tool_list else None
        description = tl.description if tl else entry["description"]
        pn_list     = tl.all_part_number_strings() if tl else entry["part_numbers"]

        pn_text = "\n".join(pn_list) if pn_list else "—"

        if has_stock:
            # Build annotated part number text
            parsed_pns  = _parse_part_numbers(pn_list)
            result: StockResult = stock_checker.check(
                type('T', (), {'part_numbers': pn_list})())
            matched_pns = {r.part_number.upper() for r in result.records}
            annotated = []
            for raw, (pn, is_opt) in zip(pn_list, parsed_pns):
                prefix = "Opt: " if is_opt else ""
                annotated.append(f"{'✔' if pn.upper() in matched_pns else ' '} {prefix}{pn}")
            pn_display = "\n".join(annotated) if annotated else "—"

            if result.found:
                for i, rec in enumerate(result.records):
                    opt_label = " (Opt)" if rec.is_opt else ""
                    values = [
                        ref_id if i == 0 else "",
                        description if i == 0 else "",
                        pn_display if i == 0 else "",
                        titles if i == 0 else "",
                        count if i == 0 else "",
                        f"{rec.part_number}{opt_label}",
                        rec.station, rec.store, rec.location, rec.qty,
                    ]
                    for col, val in enumerate(values, 1):
                        _dat(ws.cell(row=row_idx, column=col, value=val),
                             bg=STOCK_OK_BG, bold=(col == 1 and i == 0))
                    row_idx += 1
            else:
                values = [ref_id, description, pn_display, titles, count,
                          "NOT FOUND", "—", "—", "—", 0]
                for col, val in enumerate(values, 1):
                    cell = ws.cell(row=row_idx, column=col, value=val)
                    _dat(cell, bg=STOCK_NO_BG, bold=(col == 1))
                    if col == 6:
                        cell.font = Font(name="Arial", size=9, bold=True, color=NOT_FOUND_FG)
                row_idx += 1
        else:
            bg = ALT_BG if row_idx % 2 == 0 else None
            values = [ref_id, description, pn_text, titles, count]
            for col, val in enumerate(values, 1):
                _dat(ws.cell(row=row_idx, column=col, value=val),
                     bg=bg, bold=(col == 1))
            row_idx += 1

    # Missing DMC notice
    if missing_dmcs:
        nr = ws.max_row + 2
        cell = ws.cell(row=nr, column=1,
                       value="⚠ Tasks NOT FOUND in zip — tools excluded above:")
        cell.font = Font(name="Arial", bold=True, color=NOT_FOUND_FG, size=9)
        ws.merge_cells(start_row=nr, start_column=1,
                       end_row=nr, end_column=len(headers))
        for i, dmc in enumerate(missing_dmcs, 1):
            c = ws.cell(row=nr + i, column=1, value=dmc)
            c.font = Font(name="Arial", color=NOT_FOUND_FG, size=9)
            c.fill = PatternFill("solid", start_color=NOT_FOUND_BG)

    if has_stock:
        _widths(ws, {"A": 14, "B": 40, "C": 42, "D": 40,
                     "E": 10, "F": 20, "G": 10, "H": 10, "I": 14, "J": 8})
    else:
        _widths(ws, {"A": 14, "B": 40, "C": 42, "D": 40, "E": 10})


# ── Sheet 2: Consumable Materials ─────────────────────────────────────────────

def _write_consumables(wb, all_tasks, consumable_list=None):
    ws = wb.create_sheet("Consumable Materials")
    ws.freeze_panes = "A2"

    headers = ["Reference ID", "Description", "Specification",
               "Material", "Supplier", "Found In Tasks"]
    for col, h in enumerate(headers, 1):
        _hdr(ws.cell(row=1, column=col, value=h), bg=CONS_HDR_BG)
    ws.row_dimensions[1].height = 28

    # Dedup by reference_id
    dedup: Dict[str, dict] = {}
    for task in all_tasks:
        if task.title == NOT_FOUND_MARKER:
            continue
        for cons in task.consumables:
            rid = cons.reference_id
            if rid not in dedup:
                dedup[rid] = {
                    "description":   cons.description,
                    "specification": cons.specification,
                    "task_titles":   [],
                }
            if cons.source_title not in dedup[rid]["task_titles"]:
                dedup[rid]["task_titles"].append(cons.source_title)

    for row_idx, (ref_id, entry) in enumerate(sorted(dedup.items()), 2):
        bg = ALT_BG if row_idx % 2 == 0 else None

        # Enrich from consumable list if available
        cl = consumable_list.lookup(ref_id) if consumable_list else None
        description   = cl.description   if cl else entry["description"]
        specification = cl.specification if cl and cl.specification else entry["specification"]
        material      = cl.material      if cl else ""
        supplier      = cl.supplier      if cl else ""

        values = [ref_id, description, specification, material, supplier,
                  "\n".join(entry["task_titles"])]
        for col, val in enumerate(values, 1):
            _dat(ws.cell(row=row_idx, column=col, value=val),
                 bg=bg, bold=(col == 1))

    _widths(ws, {"A": 12, "B": 50, "C": 30, "D": 35, "E": 10, "F": 45})


# ── Sheet 3: Expendables/Parts ─────────────────────────────────────────────────

def _write_expendables(wb, all_tasks):
    ws = wb.create_sheet("Expendables-Parts")
    ws.freeze_panes = "A2"

    headers = ["Source Task", "AMM Item Description",
               "IPD Figure Title", "IPD DMC",
               "IPD Item #", "Part Number", "Part Description", "Qty"]
    for col, h in enumerate(headers, 1):
        _hdr(ws.cell(row=1, column=col, value=h), bg=EXP_HDR_BG)
    ws.row_dimensions[1].height = 28

    row_idx = 2
    for task in all_tasks:
        if task.title == NOT_FOUND_MARKER or not task.expendables:
            continue
        for exp in task.expendables:
            bg = EXP_ALT_BG if row_idx % 2 == 0 else EXP_ROW_BG
            pn_found = bool(exp.part_number)
            values = [
                task.title,
                exp.amm_item,
                exp.ipd_figure_title,
                exp.ipd_dmc,
                exp.ipd_item_number,
                exp.part_number   if pn_found else "NOT RESOLVED",
                exp.part_description,
                exp.quantity,
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                is_error = col == 6 and not pn_found
                _dat(cell, bg=NOT_FOUND_BG if is_error else bg,
                     bold=(col == 6),
                     color=NOT_FOUND_FG if is_error else "000000")
            row_idx += 1

    _widths(ws, {"A": 40, "B": 25, "C": 40, "D": 35,
                 "E": 10, "F": 22, "G": 35, "H": 8})